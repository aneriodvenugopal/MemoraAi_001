"""
file_ingestion — pull a remote file (or local path) and convert to plain
text (or pass-through native PDFs) so we can push it into Gemini File
Search.

Supported types (auto-detected by mime/suffix):
  • text / html / md / json / xml / csv  → read-as-text
  • pdf                                  → native upload (Gemini parses it)
  • docx                                 → python-docx → text
  • xlsx                                 → openpyxl    → text
  • images (png/jpg/jpeg/webp)           → Gemini Vision OCR → text

Audio / video transcription is intentionally NOT done here (P2 backlog).

Public API:
    fetch_to_temp(url) -> (local_path, mime, suffix)   # downloads or returns None
    extract_text(path, mime) -> str                    # converts to UTF-8 text or "" if image OCR fails
    is_native_pdf(mime, suffix) -> bool                # true → upload as-is to Gemini
"""
from __future__ import annotations
import logging
import os
import tempfile
from typing import Optional, Tuple
from urllib.parse import urlparse

import httpx

logger = logging.getLogger(__name__)

NATIVE_PDF_SUFFIXES = (".pdf",)
TEXT_SUFFIXES = (".txt", ".md", ".markdown", ".html", ".htm",
                 ".json", ".xml", ".csv", ".tsv", ".log")
DOCX_SUFFIXES = (".docx",)
XLSX_SUFFIXES = (".xlsx", ".xlsm")
IMAGE_SUFFIXES = (".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp")
SKIP_SUFFIXES = (".mp3", ".wav", ".m4a", ".mp4", ".mov", ".avi", ".mkv",
                 ".aac", ".flac", ".ogg")

DOWNLOAD_MAX_BYTES = 25 * 1024 * 1024  # 25 MB


def _infer_suffix(url: str, mime: str) -> str:
    parsed = urlparse(url)
    path = parsed.path or ""
    _, ext = os.path.splitext(path.lower())
    if ext:
        return ext
    if "pdf" in mime:
        return ".pdf"
    if "wordprocessingml" in mime or "msword" in mime:
        return ".docx"
    if "spreadsheetml" in mime or "excel" in mime:
        return ".xlsx"
    if "image/jpeg" in mime:
        return ".jpg"
    if "image/png" in mime:
        return ".png"
    if "image/webp" in mime:
        return ".webp"
    return ".bin"


async def fetch_to_temp(url: str) -> Optional[Tuple[str, str, str]]:
    """Download `url` to a tempfile. Returns (path, mime, suffix) or None."""
    if not url or not url.startswith(("http://", "https://")):
        return None
    try:
        async with httpx.AsyncClient(
            timeout=30.0, follow_redirects=True
        ) as client:
            async with client.stream("GET", url) as r:
                if r.status_code >= 400:
                    return None
                mime = (r.headers.get("content-type") or "").split(";")[0].strip()
                suffix = _infer_suffix(url, mime)
                if suffix in SKIP_SUFFIXES:
                    return None
                fd, path = tempfile.mkstemp(suffix=suffix)
                size = 0
                with os.fdopen(fd, "wb") as f:
                    async for chunk in r.aiter_bytes():
                        size += len(chunk)
                        if size > DOWNLOAD_MAX_BYTES:
                            f.close()
                            os.unlink(path)
                            return None
                        f.write(chunk)
                return path, mime, suffix
    except Exception as e:  # noqa: BLE001
        logger.warning(f"fetch_to_temp({url}) failed: {e}")
        return None


def is_native_pdf(mime: str, suffix: str) -> bool:
    return ("pdf" in (mime or "")) or (suffix in NATIVE_PDF_SUFFIXES)


def is_image(mime: str, suffix: str) -> bool:
    return (mime or "").startswith("image/") or suffix in IMAGE_SUFFIXES


def extract_text(path: str, mime: str = "") -> str:
    """Best-effort plain-text extraction. Returns "" if not extractable."""
    if not path or not os.path.exists(path):
        return ""
    suffix = os.path.splitext(path.lower())[1]

    if suffix in TEXT_SUFFIXES:
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()
        except Exception as e:  # noqa: BLE001
            logger.warning(f"text read failed: {e}")
            return ""

    if suffix in DOCX_SUFFIXES:
        try:
            from docx import Document
            doc = Document(path)
            return "\n".join(p.text for p in doc.paragraphs if p.text)
        except Exception as e:  # noqa: BLE001
            logger.warning(f"docx extract failed: {e}")
            return ""

    if suffix in XLSX_SUFFIXES:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            lines = []
            for ws in wb.worksheets:
                lines.append(f"## Sheet: {ws.title}")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) if c is not None else "" for c in row]
                    if any(cells):
                        lines.append("\t".join(cells))
            return "\n".join(lines)
        except Exception as e:  # noqa: BLE001
            logger.warning(f"xlsx extract failed: {e}")
            return ""

    if suffix in NATIVE_PDF_SUFFIXES:
        # Best-effort fallback for callers that don't use native upload.
        try:
            from pypdf import PdfReader
            reader = PdfReader(path)
            return "\n\n".join(
                (p.extract_text() or "") for p in reader.pages
            )
        except Exception as e:  # noqa: BLE001
            logger.warning(f"pdf extract failed: {e}")
            return ""

    if suffix in IMAGE_SUFFIXES or (mime or "").startswith("image/"):
        return _ocr_image_via_gemini(path)

    return ""


def _ocr_image_via_gemini(path: str) -> str:
    """Run image → text OCR using Gemini vision. Returns "" on failure."""
    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        return ""
    try:
        from google import genai
        from google.genai import types
        client = genai.Client(api_key=api_key)
        with open(path, "rb") as f:
            data = f.read()
        suffix = os.path.splitext(path.lower())[1]
        mime = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                "webp": "image/webp", "gif": "image/gif", "bmp": "image/bmp"}.get(
            suffix.lstrip("."), "image/jpeg")
        prompt = (
            "Extract ALL text visible in this image verbatim — including "
            "headings, prices, numbers, RERA codes, contact info, addresses, "
            "tables. Preserve numbers EXACTLY. Return plain text only."
        )
        resp = client.models.generate_content(
            model="gemini-2.5-flash",
            contents=[
                types.Content(role="user", parts=[
                    types.Part(inline_data=types.Blob(mime_type=mime, data=data)),
                    types.Part(text=prompt),
                ]),
            ],
            config=types.GenerateContentConfig(
                temperature=0.0, max_output_tokens=2000,
                thinking_config=types.ThinkingConfig(thinking_budget=0),
            ),
        )
        return (resp.text or "").strip()
    except Exception as e:  # noqa: BLE001
        logger.warning(f"image OCR failed: {e}")
        return ""
